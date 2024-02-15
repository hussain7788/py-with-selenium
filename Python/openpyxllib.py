import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('output.xlsx')
import pdb;pdb.set_trace()
# Select a sheet
sheet = workbook['My Sheet']

# Access cell values
cell_value = sheet['A1'].value
print(f'Value at A1: {cell_value}')

# Iterate through rows
for row in sheet.iter_rows(min_row=1, max_row=3, values_only=True):
    print(row)



#### write inot xl file
import openpyxl

# Create a new workbook
workbook = openpyxl.Workbook()

# Select a sheet
sheet = workbook.active

# Write values to cells
sheet['A1'] = 'Hello'
sheet['B1'] = 'World'

# Save the workbook
workbook.save('example_write.xlsx')


#### creating charts

from openpyxl.chart import BarChart, Reference

# Assume we have a sheet with data in columns A and B

# Create a new workbook
workbook = openpyxl.Workbook()
sheet = workbook.active

# Add data to the sheet
data = [
    ('Category', 'Value'),
    ('A', 10),
    ('B', 20),
    ('C', 15),
]

for row in data:
    sheet.append(row)

# Create a bar chart
chart = BarChart()
values = Reference(sheet, min_col=2, min_row=1, max_col=2, max_row=len(data))
categories = Reference(sheet, min_col=1, min_row=2, max_row=len(data))

chart.add_data(values, titles_from_data=True)
chart.set_categories(categories)

# Add the chart to the sheet
sheet.add_chart(chart, "D5")

# Save the workbook
workbook.save('example_chart.xlsx')
